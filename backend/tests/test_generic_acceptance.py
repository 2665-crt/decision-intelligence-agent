from __future__ import annotations

from hashlib import sha256
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from studio_api.execution import ComputedFinding, ExecutionResult, FindingEvidence, execute_plan
from studio_api.planning import build_plan
from studio_api.profiling import profile_file, profile_files, read_tables
from studio_api.relationships import discover_relationships
from studio_api.validation import validate_result


def _write_csv(tmp_path: Path, name: str, frame: pd.DataFrame) -> Path:
    source = tmp_path / name
    frame.to_csv(source, index=False)
    return source


def _run_pipeline(source: Path, question: str):
    profile = profile_file(source)
    plan = build_plan(profile, question)
    execution = execute_plan(dict(read_tables(source)), plan)
    validated = validate_result(execution, profile)
    return profile, plan, execution, validated


def _assert_traceable_success(profile, plan, execution, validated, calculation: str) -> None:
    assert plan.status == "READY"
    assert execution.status == "SUCCESS"
    assert validated.status == "SUCCESS"
    assert validated.findings
    assert validated.answer == "；".join(finding["conclusion"] for finding in validated.findings)
    for finding in validated.findings:
        evidence = finding["evidence"]
        assert evidence["source"] == {"file_hash": profile.file_hash, "table": plan.table}
        assert evidence["calculation"] == calculation
        assert evidence["fields"]
        assert evidence["row_indices"]
        expected_output = (
            {evidence["grouping"][0]: finding["value"], "aggregate": finding["metric_value"]}
            if finding["kind"] == "ranking" and evidence["grouping"] and finding["metric_value"] is not None
            else finding["value"]
        )
        assert evidence["output_value"] == expected_output
        assert evidence["metric_value"] == finding["metric_value"]


def test_ecommerce_product_sales_ranking_runs_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "ecommerce.csv",
        pd.DataFrame(
            {
                "product_name": ["A", "B", "A", "B"],
                "sales_amount": [120, 110, 30, 20],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "哪个产品销售额最高？")

    assert plan.fields == {"dimension": "product_name", "metric": "sales_amount"}
    assert validated.findings[0]["value"] == "A"
    assert validated.findings[0]["metric_value"] == 150.0
    _assert_traceable_success(profile, plan, execution, validated, "groupby(product_name).sum(sales_amount)")


def test_ranking_uses_the_explicit_chinese_product_dimension_not_an_unrequested_region(tmp_path):
    source = _write_csv(
        tmp_path,
        "sales.csv",
        pd.DataFrame(
            {
                "商品": ["A", "B", "C", "D"],
                "销量": [10, 25, 5, 20],
                "销售额": [100, 375, 80, 260],
                "地区": ["华东", "华北", "华东", "华南"],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "按销量对商品进行排名，并给出最高商品和数据证据")

    assert plan.fields == {"dimension": "商品", "metric": "销量"}
    assert execution.status == "SUCCESS"
    assert validated.findings[0]["value"] == "B"
    assert validated.findings[0]["metric_value"] == 25.0
    assert "商品 中的 B" in validated.answer


def test_traffic_congestion_comparison_runs_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "traffic.csv",
        pd.DataFrame(
            {
                "road_segment": ["east", "east", "west", "west"],
                "congestion_index": [0.8, 0.9, 0.6, 0.7],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(
        source, "比较 road_segment 的 congestion_index 差异"
    )

    assert plan.operations == ("group_comparison",)
    values = validated.findings[0]["value"]
    assert [item["group"] for item in values] == ["east", "west"]
    assert [item["value"] for item in values] == pytest.approx([0.85, 0.65])
    _assert_traceable_success(
        profile,
        plan,
        execution,
        validated,
        "groupby(road_segment).mean(congestion_index)",
    )


def test_subject_score_difference_runs_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "scores.csv",
        pd.DataFrame(
            {
                "subject": ["math", "math", "language", "language"],
                "score_value": [88, 92, 68, 72],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "比较 subject 的 score_value 差异")

    assert validated.findings[0]["metric_value"] == 90.0
    assert validated.findings[0]["value"][0] == {"group": "math", "value": 90.0}
    _assert_traceable_success(profile, plan, execution, validated, "groupby(subject).mean(score_value)")


def test_server_error_rate_anomaly_runs_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "server-log.csv",
        pd.DataFrame(
            {
                "observed_at": pd.date_range("2025-01-01", periods=5, freq="h").astype(str),
                "error_rate": [0.01, 0.02, 0.015, 0.018, 0.95],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "error_rate 有哪些异常值？")

    assert plan.fields == {"metric": "error_rate", "time": "observed_at"}
    assert validated.findings[0]["value"] == 0.95
    assert validated.findings[0]["evidence"]["row_indices"] == [4]
    _assert_traceable_success(profile, plan, execution, validated, "iqr_outliers(error_rate, k=1.5)")


def test_inventory_risk_is_conservatively_insufficient_when_no_operator_matches(tmp_path):
    source = _write_csv(
        tmp_path,
        "inventory.csv",
        pd.DataFrame(
            {
                "sku": ["A", "B", "C", "D"],
                "stock_on_hand": [2, 80, 40, 15],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "识别低库存风险")

    stock = next(column for column in profile.tables[0].columns if column.name == "stock_on_hand")
    assert stock.semantic_role == "metric" and stock.confidence >= 0.70
    assert plan.status == "INSUFFICIENT_DATA"
    assert execution.status == "INSUFFICIENT_DATA"
    assert validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == ()
    assert validated.evidence == ()
    assert "可执行的排名、趋势、异常、分组比较、相关性或预测算子" in validated.answer


def test_financial_decline_uses_nonstandard_fields_through_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "finance-generic.csv",
        pd.DataFrame(
            {
                "period_key": ["2025-01-01", "2025-02-01", "2025-03-01"],
                "net_result_x": [120, 100, 80],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(
        source, "net_result_x 随 period_key 的趋势如何？"
    )

    assert plan.fields == {"time": "period_key", "metric": "net_result_x"}
    assert validated.findings[0]["value"][-1] == {"time": "2025-03-01", "value": 80.0}
    assert "从 120 变化到 80" in validated.answer
    assert all(fixed_name not in plan.fields.values() for fixed_name in ("营业收入", "营业利润"))
    _assert_traceable_success(
        profile,
        plan,
        execution,
        validated,
        "groupby(period_key).sum(net_result_x).sort_index()",
    )


def test_forecast_without_time_is_rejected_by_the_full_pipeline(tmp_path):
    source = _write_csv(
        tmp_path,
        "forecast-without-time.csv",
        pd.DataFrame({"bucket": ["A", "B", "A"], "measure_x": [10, 12, 14]}),
    )

    profile, plan, execution, validated = _run_pipeline(source, "预测 measure_x 未来趋势")

    assert profile.tables[0].row_count == 3
    assert plan.operations == ("forecast",)
    assert plan.fields == {"metric": "measure_x"}
    assert plan.status == execution.status == validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == () and validated.evidence == ()
    assert any("可信的时间字段" in limitation for limitation in validated.limitations)


def test_nonstandard_field_names_answer_only_from_profile_semantics(tmp_path):
    source = _write_csv(
        tmp_path,
        "nonstandard.csv",
        pd.DataFrame(
            {
                "dt": ["2025-01-01", "2025-01-02", "2025-01-03", "2025-01-04"],
                "col1": ["alpha", "beta", "alpha", "beta"],
                "value_x": [10, 11, 15, 12],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "哪个 col1 的 value_x 最高？")
    columns = {column.name: column for column in profile.tables[0].columns}

    assert (columns["dt"].semantic_role, columns["col1"].semantic_role, columns["value_x"].semantic_role) == (
        "time",
        "dimension",
        "metric",
    )
    assert min(columns[name].confidence for name in ("dt", "col1", "value_x")) >= 0.70
    assert validated.findings[0]["value"] == "alpha"
    assert validated.findings[0]["metric_value"] == 25.0
    _assert_traceable_success(profile, plan, execution, validated, "groupby(col1).sum(value_x)")


def test_multi_sheet_only_high_confidence_relationships_are_auto_usable(tmp_path):
    source = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    customers = workbook.active
    customers.title = "customers"
    customers.append(["customer_id", "segment"])
    for row in ((101, "A"), (102, "B"), (103, "A"), (104, "B")):
        customers.append(row)
    orders = workbook.create_sheet("orders")
    orders.append(["customer_id", "region", "load_value"])
    for row in ((101, "north", 10), (102, "south", 20), (103, "north", 30), (104, "south", 5)):
        orders.append(row)
    activity = workbook.create_sheet("activity")
    activity.append(["customer_id", "event_type"])
    for row in ((101, "open"), (101, "click"), (105, "open"), (105, "click")):
        activity.append(row)
    workbook.save(source)

    profile, plan, execution, validated = _run_pipeline(source, "哪个 region 的 load_value 最高？")
    relationships = discover_relationships([profile])
    auto = [candidate for candidate in relationships if candidate.can_auto_use]
    guarded = [candidate for candidate in relationships if not candidate.can_auto_use]

    assert len(auto) == 1
    assert {auto[0].left_table, auto[0].right_table} == {"customers", "orders"}
    assert auto[0].confidence >= 0.85 and auto[0].requires_confirmation is False
    assert guarded and all(candidate.requires_confirmation for candidate in guarded)
    assert all(candidate.confidence < 0.85 for candidate in guarded)
    assert plan.table == "orders"
    assert set(plan.fields.values()) == {"region", "load_value"}
    assert all(candidate.left_field not in plan.fields.values() for candidate in guarded)
    _assert_traceable_success(profile, plan, execution, validated, "groupby(region).sum(load_value)")


def test_multi_file_profiles_and_relationships_are_reproducible_and_hashed(tmp_path):
    customers = _write_csv(
        tmp_path,
        "customers.csv",
        pd.DataFrame({"customer_id": [101, 102, 103, 104], "segment": ["A", "B", "A", "B"]}),
    )
    orders = _write_csv(
        tmp_path,
        "orders.csv",
        pd.DataFrame(
            {
                "customer_id": [101, 102, 103, 104],
                "region": ["north", "south", "north", "south"],
                "order_value": [10, 20, 30, 5],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(orders, "哪个 region 的 order_value 最高？")
    first_profiles = profile_files([customers, orders])
    second_profiles = profile_files([customers, orders])
    first_relationships = discover_relationships(first_profiles)
    second_relationships = discover_relationships(second_profiles)

    assert [item.to_dict() for item in first_profiles] == [item.to_dict() for item in second_profiles]
    assert [item.to_dict() for item in first_relationships] == [item.to_dict() for item in second_relationships]
    assert {item.file_hash for item in first_profiles} == {
        sha256(customers.read_bytes()).hexdigest(),
        sha256(orders.read_bytes()).hexdigest(),
    }
    assert len(first_relationships) == 1
    relationship = first_relationships[0]
    assert relationship.left_file_hash and relationship.right_file_hash
    assert relationship.left_file_hash != relationship.right_file_hash
    assert relationship.can_auto_use is True
    _assert_traceable_success(profile, plan, execution, validated, "groupby(region).sum(order_value)")


def test_unknown_low_confidence_fields_never_produce_a_number(tmp_path):
    source = _write_csv(
        tmp_path,
        "one-row.csv",
        pd.DataFrame({"mystery_label": ["A"], "mystery_value": [999999]}),
    )

    profile, plan, execution, validated = _run_pipeline(
        source, "哪个 mystery_label 的 mystery_value 最高？"
    )
    columns = profile.tables[0].columns

    assert all(column.confidence < 0.70 for column in columns)
    assert plan.fields == {}
    assert plan.status == execution.status == validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == () and validated.evidence == ()
    assert "999999" not in validated.answer


def test_sales_amount_never_answers_a_sales_volume_question(tmp_path):
    source = _write_csv(
        tmp_path,
        "sales-amount-only.csv",
        pd.DataFrame(
            {
                "product_name": ["A", "B", "A", "B"],
                "sales_amount": [12, 20, 15, 1],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "哪个产品销量最高？")

    assert plan.fields == {"dimension": "product_name"}
    assert "sales_amount" not in plan.fields.values()
    assert plan.status == execution.status == validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == () and validated.evidence == ()
    assert profile.file_hash


def test_non_numeric_and_nan_rows_do_not_pollute_ranking_evidence(tmp_path):
    source = _write_csv(
        tmp_path,
        "dirty-values.csv",
        pd.DataFrame(
            {
                "product_name": ["A", "A", "A", "A", "A", "A", "B", "B", "B", "B", "B", "B"],
                "sales_amount": [12, "bad", 15, None, 0, 0, 20, 1, 1, 1, 1, 1],
            }
        ),
    )

    profile, plan, execution, validated = _run_pipeline(source, "哪个产品销售额最高？")

    assert validated.findings[0]["value"] == "A"
    assert validated.findings[0]["metric_value"] == 27.0
    assert validated.findings[0]["evidence"]["row_indices"] == [0, 2, 4, 5]
    assert 1 not in validated.findings[0]["evidence"]["row_indices"]
    assert 3 not in validated.findings[0]["evidence"]["row_indices"]
    _assert_traceable_success(profile, plan, execution, validated, "groupby(product_name).sum(sales_amount)")


def test_unsafe_identifier_and_amount_relationships_are_not_auto_used(tmp_path):
    accounts = _write_csv(tmp_path, "accounts.csv", pd.DataFrame({"account_id": [101, 102, 103]}))
    groups = _write_csv(tmp_path, "groups.csv", pd.DataFrame({"group_id": [101, 102, 103]}))
    first_paid = _write_csv(tmp_path, "paid-a.csv", pd.DataFrame({"paid_amount": [10, 20, 30]}))
    second_paid = _write_csv(tmp_path, "paid-b.csv", pd.DataFrame({"paid_amount": [10, 20, 30]}))

    account_profile, plan, execution, validated = _run_pipeline(accounts, "account_id 有哪些异常值？")
    relationships = discover_relationships(profile_files([accounts, groups, first_paid, second_paid]))

    assert plan.status == execution.status == validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == () and validated.evidence == ()
    assert not any({candidate.left_field, candidate.right_field} == {"account_id", "group_id"} for candidate in relationships)
    paid_candidates = [
        candidate
        for candidate in relationships
        if candidate.left_field == candidate.right_field == "paid_amount"
    ]
    assert paid_candidates
    assert all(not candidate.can_auto_use and candidate.requires_confirmation for candidate in paid_candidates)
    assert account_profile.file_hash


@pytest.mark.parametrize(
    ("value", "calculation", "expected_limitation"),
    [
        (20.0, "", "calculation"),
        (None, "groupby(bucket).sum(measure_x)", "output value"),
    ],
)
def test_validator_rejects_numeric_findings_without_calculation_or_value(
    tmp_path, value, calculation, expected_limitation
):
    source = _write_csv(
        tmp_path,
        "invalid-finding.csv",
        pd.DataFrame({"bucket": ["A", "B"], "measure_x": [12, 20]}),
    )
    profile = profile_file(source)
    invalid = ComputedFinding(
        kind="ranking",
        value=value,
        metric_value=20.0,
        conclusion="B 的 measure_x 为 20。",
        confidence=0.8,
        evidence=FindingEvidence(
            source={"file_hash": profile.file_hash, "table": profile.tables[0].name},
            fields=("bucket", "measure_x"),
            filters=(),
            grouping=("bucket",),
            calculation=calculation,
            row_indices=(1,),
        ),
    )

    validated = validate_result(ExecutionResult("SUCCESS", (invalid,), ()), profile)

    assert validated.status == "INSUFFICIENT_DATA"
    assert validated.findings == () and validated.evidence == ()
    assert any(expected_limitation in limitation for limitation in validated.limitations)


def test_validated_answer_contains_no_numeric_claim_outside_computed_findings(tmp_path):
    source = _write_csv(
        tmp_path,
        "traceability.csv",
        pd.DataFrame({"bucket": ["A", "B", "A", "B"], "measure_x": [12, 20, 15, 1]}),
    )

    profile, plan, execution, validated = _run_pipeline(source, "哪个 bucket 的 measure_x 最高？")

    assert validated.answer == "bucket 中的 A 对应 measure_x 汇总值为 27。"
    assert validated.answer == validated.findings[0]["conclusion"]
    assert validated.findings[0]["metric_value"] == 27.0
    assert validated.findings[0]["evidence"]["metric_value"] == 27.0
    assert validated.findings[0]["evidence"]["output_value"] == {"bucket": "A", "aggregate": 27.0}
    _assert_traceable_success(profile, plan, execution, validated, "groupby(bucket).sum(measure_x)")
